import openpyxl


class HomePageData:
    test_HomePage_data = [{"firstname":"sai","lastname": "krishna","gender": "Male"},{"firstname":"komi","lastname": "kondra","gender": "Female"}]

    @staticmethod
    def getTestData(testcase_name):
        Dict = { }
        book = openpyxl.load_workbook("C:\\Users\\komali\\OneDrive\\Desktop\\Datasets\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == testcase_name:
                for j in range(1, sheet.max_column + 1):  # to get columns
                    # for j in range(2,sheet.max_column+1):      #to get only from column 2
                    # print(sheet.cell(row=i , column=j).value)
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]
